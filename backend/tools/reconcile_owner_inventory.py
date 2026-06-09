from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from sqlalchemy import text

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend" / "src"))

from openpyxl import load_workbook  # noqa: E402

from ai_inventory_backend.repository import InventoryRepository  # noqa: E402


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare an exported inventory workbook with the latest DB batch.")
    parser.add_argument("excel_path", type=Path, help="Path to the exported inventory workbook.")
    parser.add_argument("--owner", required=True, help="Owner name used to filter the database rows.")
    args = parser.parse_args()

    excel_rows = _read_excel_spu_totals(args.excel_path)
    db_rows = _read_db_spu_totals(args.owner)

    excel_by_spu = {str(row["spu"]): row for row in excel_rows if row.get("spu")}
    db_by_spu = {str(row["spu"]): row for row in db_rows if row.get("spu")}
    all_spus = sorted(set(excel_by_spu) | set(db_by_spu))

    diffs = []
    for spu in all_spus:
        excel_total = _num(excel_by_spu.get(spu, {}).get("total_inventory"))
        db_total = _num(db_by_spu.get(spu, {}).get("total_inventory"))
        diffs.append(
            {
                "spu": spu,
                "excel_total": excel_total,
                "db_total": db_total,
                "diff_db_minus_excel": round(db_total - excel_total, 2),
                "excel_present": spu in excel_by_spu,
                "db_present": spu in db_by_spu,
                "db_sku_count": db_by_spu.get(spu, {}).get("sku_count", 0),
                "db_warning": db_by_spu.get(spu, {}).get("warning_summary", ""),
            }
        )

    diffs.sort(key=lambda row: abs(row["diff_db_minus_excel"]), reverse=True)
    result = {
        "excel_total": round(sum(_num(row.get("total_inventory")) for row in excel_rows), 2),
        "db_total": round(sum(_num(row.get("total_inventory")) for row in db_rows), 2),
        "diff_db_minus_excel": round(
            sum(_num(row.get("total_inventory")) for row in db_rows)
            - sum(_num(row.get("total_inventory")) for row in excel_rows),
            2,
        ),
        "excel_spu_count": len(excel_by_spu),
        "db_spu_count": len(db_by_spu),
        "top_diffs": diffs[:20],
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


def _read_excel_spu_totals(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    spu_idx = _find_idx(headers, ["维度", "SPU"])
    total_idx = _find_idx(headers, ["总库存"])
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        spu = values[spu_idx] if spu_idx < len(values) else None
        total = values[total_idx] if total_idx < len(values) else None
        if spu in (None, "", "合计"):
            continue
        rows.append({"spu": str(spu).strip(), "total_inventory": _num(total)})
    return rows


def _read_db_spu_totals(owner: str) -> list[dict]:
    repo = InventoryRepository()
    batch = repo._latest_batch()
    sql = text(
        """
        select spu,
               count(*) as sku_count,
               sum(coalesce("总库存", 0)) as total_inventory,
               string_agg(distinct coalesce("备货预警", '未标记'), ',') as warning_summary
        from lx_ads.ads_lx_kd_inventory_sku_calc
        where insert_time = :insert_time
          and "归属" = :owner
        group by spu
        order by total_inventory desc
        """
    )
    with repo.engine.connect() as conn:
        rows = conn.execute(sql, {"insert_time": batch["insert_time"], "owner": owner}).mappings().all()
    return [dict(row) for row in rows]


def _find_idx(headers: list[str], candidates: list[str]) -> int:
    for candidate in candidates:
        for index, header in enumerate(headers):
            if candidate == header or candidate in header:
                return index
    raise RuntimeError(f"Cannot find any header in {candidates}; headers={headers}")


def _num(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


if __name__ == "__main__":
    raise SystemExit(main())
